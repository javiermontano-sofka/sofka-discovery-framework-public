#!/usr/bin/env python3
"""XLSX → priming-rag. Per-sheet schema + head + named ranges."""
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from _common import cli, file_meta, write_priming, truncate


def main() -> None:
    args = cli()
    path = Path(args.path)
    wb = load_workbook(path, data_only=True, read_only=True)
    sheets = wb.sheetnames
    named_ranges = [d.name for d in wb.defined_names.definedName] if hasattr(wb.defined_names, 'definedName') else list(wb.defined_names)
    meta = file_meta(path)

    blocks: list[str] = []
    evidencia: list[str] = []
    for sh in sheets[:20]:
        try:
            df = pd.read_excel(path, sheet_name=sh, nrows=500)
        except Exception as e:
            blocks.append(f"### Sheet: {sh}\n_error: {e}_")
            continue
        cols = ", ".join(map(str, df.columns[:12]))
        head = df.head(15).to_markdown(index=False)
        blocks.append(f"### Sheet: `{sh}` ({len(df)} rows × {len(df.columns)} cols)\n\ncols: {cols}\n\n{head}")
        evidencia.append(f"[ADJUNTO:{path.name}:sheet={sh}]")

    resumen = [
        f"Workbook con {len(sheets)} hojas: {', '.join(sheets[:10])}",
        f"Named ranges: {len(named_ranges)}" + (f" ({', '.join(named_ranges[:5])}...)" if named_ranges else ""),
    ]
    contenido = "\n\n".join(blocks)
    if named_ranges:
        contenido += "\n\n### Named ranges\n" + "\n".join(f"- `{n}`" for n in named_ranges[:30])
    out = write_priming(Path(args.out) if args.out else None, meta, "xlsx", resumen, truncate(contenido), evidencia)
    print(str(out))


if __name__ == "__main__":
    main()
